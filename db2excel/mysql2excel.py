# coding:utf8  
import sys  
import openpyxl
#import MySQLdb
import pymysql as MySQLdb
import datetime  
import time 
import logging
import yaml
import traceback

f = open('Conf.yaml', encoding='utf-8')
conf = yaml.load(f, Loader=yaml.FullLoader)

print('conf---')
print(conf)

# 设置日志
LOG_FORMAT = conf['LOG']['LOG_FORMAT'] # "%(asctime)s - %(levelname)s - %(message)s"
DATE_FORMAT = conf['LOG']['DATE_FORMAT'] # "%m/%d/%Y %H:%M:%S %p"
LOG_PATH = conf['LOG']['LOG_PATH']
logging.basicConfig(filename=LOG_PATH + 'log.log', level=logging.INFO, format=LOG_FORMAT, datefmt=DATE_FORMAT)
logging.info("starting...")

# 数据库信息
host = conf['DB']['HOST'] # '172.16.0.52'  
user = conf['DB']['USER'] # 'query01'  
pwd = conf['DB']['PWD'] # 'query01'  
db = conf['DB']['DB'] # 'small_core'  
sql = conf['DB']['SQL'] # 'select * from VIEW_REPORT_PROJECT_RENT'  

# excel 信息
strToday = time.strftime("%Y-%m-%d", time.localtime(int(time.time()))) 
sheet_name = conf['EXCEL']['SHEET_NAME'] # '租金收取情况报表' 
out_path = conf['EXCEL']['OUT_PATH'] + conf['EXCEL']['FILE_BASE'] + strToday + '.xlsx'

# todo: 以上，可以修改为配置文件中读取
logging.info("Out Path: " + out_path)

# 读取数据库
logging.info("connect to database ")
logging.info("Query: " + sql)
conn = MySQLdb.connect(host, user, pwd, db, charset='utf8')  
cursor = conn.cursor()  
try:
    logging.info("read from table ")
    count = cursor.execute(sql)
    logging.info("read from table count: ")
    logging.info(count)
    # print(count)  
    cursor.scroll(0, mode='absolute')  
    results = cursor.fetchall()
    fields = cursor.description

    print('fields: ')
    print(fields)

    logging.info("prepare excel file ")
    # 写入 Excel
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.active
    # outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
    outws.title = sheet_name

    logging.info("write to excel file ")
    # print(len(fields))

    for field in range(0, len(fields)): 
        # print(fields[field][0])
        outws.cell(1, field+1, fields[field][0])
     
    i = 2  # 注意：'cell'函数中行列起始值为1
    for line in results:  
        for x in range(0,len(line)):  	     
            outws.cell(column = x+1 , row = i , value = "%s" % line[x])  
        i += 1  

    outwb.save(out_path)
    logging.info("write to excel file: done ")

except Exception as e:
    traceback.print_exc()
    logging.warning("exec failed, failed msg:" + traceback.format_exc())

logging.info("closeing database ")
conn.close
logging.info("closeing database: done ")
