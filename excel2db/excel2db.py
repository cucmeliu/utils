# coding:utf8  
import sys  
import openpyxl
from openpyxl import load_workbook
#import MySQLdb
import pymysql as MySQLdb
import datetime  
import time 
import logging
import yaml
import traceback

# pyinstaller -F excel2db.py 打包生成 exe

f = open('Conf.yaml', encoding='utf-8')
conf = yaml.load(f, Loader=yaml.FullLoader)

# 设置日志
LOG_FORMAT  = conf['LOG']['LOG_FORMAT'] # "%(asctime)s - %(levelname)s - %(message)s"
DATE_FORMAT = conf['LOG']['DATE_FORMAT'] # "%m/%d/%Y %H:%M:%S %p"
LOG_PATH    = conf['LOG']['LOG_PATH']
logging.basicConfig(filename=LOG_PATH + 'log.log', level=logging.INFO, format=LOG_FORMAT, datefmt=DATE_FORMAT)
logging.info("starting...")
# 数据库信息
host    = conf['DB']['HOST'] # '172.16.0.52'  
user    = conf['DB']['USER'] # 'query01'  
pwd     = conf['DB']['PWD'] # 'query01'  
db      = conf['DB']['DB'] # 'small_core'  
table   = conf['DB']['TABLE'] # table name
query   = conf['DB']['SQL'] 

# excel 信息
in_file     = conf['EXCEL']['FILE_PATH'] + conf['EXCEL']['FILE_NAME']
logging.info(in_file)
sheet_name  = conf['EXCEL']['SHEET_NAME'] 

wb2         = load_workbook(in_file)
ws          = wb2[sheet_name]
# 
logging.info("IMPORT FILE: " + in_file)
# 读取数据库
logging.info("connect to database ")
logging.info("Query: " + query)
# 建立一个MySQL连接
conn = MySQLdb.connect(host, user, pwd, db, charset='utf8')  
# 获得游标对象, 用于逐行遍历数据库数据
cursor = conn.cursor()  
try:
    # 先清空表
    logging.info("TRUNCATE table ")
    cursor.execute("TRUNCATE table " + table)
    conn.commit()

    logging.info("read from EXCEL ")
    list = []

    for r in ws.iter_rows(min_row=2, min_col=1):
        data=(r[0].value, r[1].value, r[2].value, r[3].value, r[4].value,
            r[5].value, r[6].value, r[7].value, r[8].value, r[9].value)
        list.append(data)
        # 单条插入，太慢了
        # values = (r[0].value, r[1].value, r[2].value, r[3].value, r[4].value,
        #     r[5].value, r[6].value, r[7].value, r[8].value, r[9].value)
        # cursor.execute(query, values)
    if (len(list) > 0):
        cursor.executemany(query, list)
        # 提交
        conn.commit()

    logging.info("transfer from excel to db: done ")

except Exception as e:
    traceback.print_exc()
    logging.warning("exec failed, failed msg:" + traceback.format_exc())

logging.info("closeing database ")
# 关闭游标
cursor.close()
# 关闭数据库连接
conn.close()
logging.info("closeing database: done ")
